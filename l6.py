import json
import os
from datetime import datetime

import requests
from openpyxl import Workbook, load_workbook


def exel_sheet():
    crypto_archive = os.path.exists("crypto_archive.xlsx")
    if not crypto_archive:
        wb = Workbook()
        wb.save("crypto_archive.xlsx")


def check_state_of_currencies():
    currencies = [
        {'id': 1, 'currency': 'btcusd'},
        {'id': 2, 'currency': 'btceur'},
        {'id': 3, 'currency': 'eurusd'},
        {'id': 4, 'currency': 'xrpusd'},
        {'id': 5, 'currency': 'xrpeur'},
        {'id': 6, 'currency': 'xrpbtc'},
        {'id': 7, 'currency': 'ltcusd'},
        {'id': 8, 'currency': 'ltceur'},
        {'id': 9, 'currency': 'ltcbtc'},
        {'id': 10, 'currency': 'ethusd'},
        {'id': 11, 'currency': 'etheur'},
        {'id': 12, 'currency': 'ethbtc'},
        {'id': 12, 'currency': 'bchusd'},
        {'id': 13, 'currency': 'bcheur'},
        {'id': 14, 'currency': 'bchbtc'},

    ]
    start = True
    while start:
        print('Jaką parę chcesz sprawdzić? Podaj numer\nDostępne:')
        for currency in currencies:
            print(f"{currency['id']}: {currency['currency']}")
        input1 = input()
        currency_from_input = currencies[int(input1) - 1]['currency']

        print("Jaką ilość posiadasz?")
        amount = float(input())

        if not any(d['currency'] == currency_from_input for d in currencies):
            print("Nie ma takiej pary, podaj inną")

        else:
            url = "https://www.bitstamp.net/api/v2/transactions/" + currency_from_input + "/"
            query = {'time': 'day'}
            response = requests.get(url, params=query)
            transactions = json.loads(response.text)
            sorted_transactions = sorted(transactions, key=lambda i: i['date'])

            last_transaction = float(sorted_transactions[0]['price'])
            transaction_24h_ago = float(sorted_transactions[-1]['price'])

            profit_percent = (last_transaction / transaction_24h_ago) * 100 - 100
            profit_amount = last_transaction * amount - transaction_24h_ago * amount
            balance_amount_abs = abs(profit_amount)

            if profit_percent >= 0:
                print("Wzrost:", "+", round(float(profit_percent), 2), "%")
                print("W ciągu ostatniej doby zarobiłeś(aś)", round(profit_amount, 2))
                balance_type = 'dodatni'
            else:
                print("Spadek:", round(float(profit_percent), 2), "%")
                print("W ciągu ostatniej doby straciłeś(aś)", round(profit_amount, 2))
                balance_type = 'ujemny'

            date = datetime.now()
            wb = load_workbook("crypto_archive.xlsx")
            print(wb.sheetnames)
            if currency_from_input not in wb.sheetnames:
                ws = wb.create_sheet(currency_from_input)
            else:
                ws = wb[currency_from_input]
            ws.cell(row=1, column=1).value = "Data"
            ws.cell(row=ws.max_row + 1, column=1).value = str(date)
            ws.cell(row=1, column=2).value = "Para"
            ws.cell(row=ws.max_row, column=2).value = str(currency_from_input)
            ws.cell(row=1, column=3).value = "Ilość środków"
            ws.cell(row=ws.max_row, column=3).value = str(amount)
            ws.cell(row=1, column=4).value = "Bilans"
            ws.cell(row=ws.max_row, column=4).value = str(balance_type)
            ws.cell(row=1, column=5).value = "Wartość bezwzględna bilansu"
            ws.cell(row=ws.max_row, column=5).value = round(balance_amount_abs, 2)
            wb.save("crypto_archive.xlsx")

            print(
                "Wpisz 'stop' jeśli chcesz zakończyć działanie programu, naciśnij dowolny klawisz by sprawdzić "
                "kolejną parę")
            condition = input()
            if condition == "stop":
                start = False


exel_sheet()
check_state_of_currencies()
